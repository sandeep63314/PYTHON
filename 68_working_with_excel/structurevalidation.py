import sys

import openpyxl
from openpyxl.styles import Font,PatternFill

ex = openpyxl.load_workbook(sys.argv[1])

rownum = 1
while ex['Sheet1']['A' + str(rownum)].value != None:
    if ex['Sheet1']['A' + str(rownum)].value == ex['Sheet1']['B' + str(rownum)].value:
        ex['Sheet1']['C' + str(rownum)].value = 'True'
        ex['Sheet1']['C' + str(rownum)].font = Font(bold = True,color = '0000FF')
        ex['Sheet1']['C' + str(rownum)].fill = PatternFill(fill_type = 'solid',start_color = '00FF00',end_color = '000000')
    else:
        ex['Sheet1']['C' + str(rownum)].value = 'False'
        ex['Sheet1']['C' + str(rownum)].font = Font(bold=True, color = '0000FF')
        ex['Sheet1']['C' + str(rownum)].fill = PatternFill(fill_type = 'solid',start_color = 'FF0000',end_color = '000000')
    rownum += 1
print(f'Structure validation completed in excel file {sys.argv[1]}')

ex.save(sys.argv[1])