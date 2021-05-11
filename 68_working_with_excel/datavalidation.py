import sys
import openpyxl
from openpyxl.styles import Font, PatternFill

wb = openpyxl.load_workbook(sys.argv[1])

if 'comparison' in wb.sheetnames:
    del wb['comparison'] # delete the 'comparison' sheet if it is already available

wb.create_sheet('comparison') # create a new sheet 'comparison'
maxrow = wb['source'].max_row # retrieve max rows in 'source' sheet
maxcol = wb['source'].max_column # retrieve max columns in 'source' sheet

for c in range(1,maxcol + 1):
    wb['comparison'].cell(1,c).value = wb['source'].cell(1,c).value # create the header in 'comparison' sheet
    wb['comparison'].cell(1,c).font = Font(bold = True)

for r in range(2, maxrow + 1):
    for c in range(1, maxcol + 1):
        if wb['source'].cell(r, c).value == wb['target'].cell(r, c).value: # compare values between 'source' sheet & 'target' sheet
            wb['comparison'].cell(r, c).value = "True"
            wb['comparison'].cell(r, c).font = Font(bold=True, color = '0000FF')
            wb['comparison'].cell(r, c).fill = PatternFill(fill_type = 'solid', start_color = '00FF00', end_color = 'FFBD00')
        else:
            wb['comparison'].cell(r, c).value = '\''+str(wb['source'].cell(r, c).value)+'\',\''+str(wb['target'].cell(r, c).value)+'\''
            wb['comparison'].cell(r, c).font = Font(bold=True, color = '0000FF')
            wb['comparison'].cell(r, c).fill = PatternFill(fill_type = 'solid', start_color = 'FF0000', end_color = 'FFBD00')

print('Data successfully validated')
wb.save(sys.argv[1])