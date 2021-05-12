import openpyxl

ex = openpyxl.load_workbook('collegeDetails.xlsx')
colnum = 65
cell = chr(colnum)
rownum = 2

# to retrieve headers
headers = [val[0].value for val in [row for row in ex['student'].columns]]
print(headers)

# to select a cell range row wise
cellrange = list(ex['student'].iter_rows(2,ex['student'].max_row,1,2))
for row in cellrange:
    print(row[0].value,row[1].value)
# to select a cell range column wise
cellrange = list(ex['student'].iter_cols(1,2,2,ex['student'].max_row))
for col in cellrange:
    print(col[0].value,col[1].value,col[2].value,col[3].value,col[4].value,col[5].value)

while (ex['student'][cell + str(rownum)].value != None):
    print(ex['student'][cell + str(rownum)].value, end=' ')
    colnum += 1
    cell = chr(colnum)

print()
ws = ex['professor']  # To set a current sheet in a workbook
print(ws['A2'].value)  # to access a cell in a worksheet
print(ws.cell(3, 1).value)  # to access a cell using rownumber and column number
